from openpyxl import load_workbook
from models import BillRecord


def read_excel_data(file_path: str) -> list[BillRecord]:
    """Read and parse Excel data dynamically based on column headers."""
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Excel file is empty")

    header_row = [str(h).strip() if h else "" for h in rows[0]]
    header_index = {name: idx for idx, name in enumerate(header_row)}

    required_columns = [
        "Parent ID",
        "Child ID",
        "Supplier",
        "Comments",
        "Invoice Amount",
        "Bank Date",
        "Tier 2 - Chart of Account",
    ]
    for col in required_columns:
        if col not in header_index:
            raise ValueError(f"Missing required column in Excel: '{col}'")

    bills: list[BillRecord] = []

    for row in rows[1:]:
        if not any(row):  # skip completely empty rows
            continue

        try:
            parent_id = str(row[header_index.get("Parent ID", "")] or "").strip()
            child_id = str(row[header_index.get("Child ID", "")] or "").strip()
            supplier = str(row[header_index.get("Supplier", "")] or "").strip()
            memo = str(row[header_index.get("Comments", "")] or "").strip()
            bank_date = str(row[header_index.get("Bank Date", "")] or "").strip()
            chart_account = str(
                row[header_index.get("Tier 2 - Chart of Account", "")] or ""
            ).strip()

            # Safely convert amount to float
            amount_str = (
                str(row[header_index.get("Invoice Amount", "")] or "")
                .replace("$", "")
                .replace(",", "")
                .strip()
            )
            amount = float(amount_str) if amount_str else 0.0

            # Determine which ID to use as record_id
            record_id = child_id if child_id else parent_id

            # Skip rows without valid IDs
            if not record_id:
                continue

            bills.append(
                BillRecord(
                    record_id=record_id,
                    supplier=supplier,
                    bank_date=bank_date,
                    chart_account=chart_account,
                    amount=amount,
                    memo=memo,
                    line_memo=memo,
                    source="excel",
                )
            )

        except Exception as e:
            print(f" Skipping row due to error: {e}")
            continue

    print(f"Loaded {len(bills)} bill records from Excel.")
    return bills
